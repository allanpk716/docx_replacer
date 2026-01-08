#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建测试用的 Excel 文件，包含各种格式的文本
"""

from openpyxl import Workbook
from openpyxl.styles import Font

def create_test_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "测试数据"

    # 测试数据
    test_data = [
        ("A1", "2x10⁹", "科学计数法（上标）"),
        ("A2", "H₂O", "水分子式（下标）"),
        ("A3", "测试粗体", "粗体文本"),
        ("A4", "测试斜体", "斜体文本"),
        ("A5", "测试下划线", "下划线文本"),
        ("A6", "混合格式文本", "包含多种格式"),
    ]

    # 注意：openpyxl 不直接支持富文本（单个单元格内多种格式）
    # 所以我们只能对整个单元格应用格式

    # 使用 Unicode 上标/下标字符
    ws['A1'] = "2x10⁹"  # ⁹ 是 Unicode 上标
    ws['A1'].font = Font(name="Calibri", size=11)

    ws['A2'] = "H₂O"  # ₂ 是 Unicode 下标
    ws['A2'].font = Font(name="Calibri", size=11)

    ws['A3'] = "测试粗体"
    ws['A3'].font = Font(bold=True, name="Calibri", size=11)

    ws['A4'] = "测试斜体"
    ws['A4'].font = Font(italic=True, name="Calibri", size=11)

    ws['A5'] = "测试下划线"
    ws['A5'].font = Font(underline="single", name="Calibri", size=11)

    ws['A6'] = "混合格式文本"
    ws['A6'].font = Font(bold=True, italic=True, name="Calibri", size=11)

    # 添加说明列
    ws['B1'] = "科学计数法（上标）"
    ws['B2'] = "水分子式（下标）"
    ws['B3'] = "粗体文本"
    ws['B4'] = "斜体文本"
    ws['B5'] = "下划线文本"
    ws['B6'] = "粗体+斜体"

    # 保存文件
    output_file = "TestFiles/FormattedTextTest.xlsx"
    wb.save(output_file)
    print(f"测试 Excel 文件已创建: {output_file}")
    print("\n测试内容:")
    print("  A1: 2x10⁹ (科学计数法，Unicode 上标)")
    print("  A2: H₂O (水分子式，Unicode 下标)")
    print("  A3: 测试粗体")
    print("  A4: 测试斜体")
    print("  A5: 测试下划线")
    print("  A6: 混合格式（粗体+斜体)")

if __name__ == "__main__":
    create_test_excel()
